#!/usr/bin/env python3
"""
gen_projetos.py — regenera web/lib/viabilidadeData.ts com TODOS os projetos.

Cada projeto de viabilidade do franqueado é uma planilha Google publicada. O link
CSV expõe só 1 aba; este script usa ?output=xlsx (workbook inteiro) e extrai, por
projeto: KPIs, séries de 5 anos, fluxo de caixa acumulado (60 meses), composição
do investimento, despesas fixas, recursos humanos e DRE anual (somando 12 meses/ano).

Uso:
    pip install openpyxl
    python scripts/gen_projetos.py
"""
import io
import os
import urllib.request

import openpyxl
from openpyxl.utils import column_index_from_string as ci

# slug, nome exibido, id de publicação
PROJETOS = [
    ("compact", "Projeto Compact", "2PACX-1vRxgJwgkxtlfhF5jXUgDw5go-APDt61uiKo3t-SeXDT2plBnfviczb-RpHkL2NQiA"),
    ("compact_plus", "Projeto Compact Plus", "2PACX-1vTlcuOq1FPXsvyzuVFj-83lLHdWF94y44tcsyLxzINDb9AHFYoeJGScm_R5872lbw"),
    ("custom", "Projeto Custom", "2PACX-1vRkpVpRv1I8yM7gLPrq_AE2Zs4DYKJUlut84IuLcfVD3eIzuOyUr2ySO4r3bA2bYA"),
    ("premium", "Projeto Premium", "2PACX-1vR2nOIVvywJxpNapLXtOd_E2_6ArSbVUX1wuOtCbLjYGAMzPnFpUo8_kEINLoBXcQ"),
    ("prime", "Projeto Prime", "2PACX-1vQPKLUtsI2kvBgbz0HG9r8fDk_Cf5EFz2lQcmfYEBEyisUXOF-G2Giv4qSmY8ZnMw"),
]
DASH = "Dashboard Franqueado (5anos)"
TIR_AM = 0.0344  # informativo (aba Premissas); fração ao mês
OUT = os.path.join(os.path.dirname(__file__), "..", "web", "lib", "viabilidadeData.ts")


def num(v):
    return v if isinstance(v, (int, float)) else 0.0


def esc(s):
    return str(s).strip().replace("\\", "").replace('"', "'")


def arr(vals, nd=2):
    return "[" + ", ".join(str(round(v, nd)) for v in vals) + "]"


def fetch(pubid):
    url = f"https://docs.google.com/spreadsheets/d/e/{pubid}/pub?output=xlsx"
    with urllib.request.urlopen(url) as r:
        return openpyxl.load_workbook(io.BytesIO(r.read()), data_only=True)


def row(ws, r, c1, c2):
    return [ws.cell(r, c).value for c in range(ci(c1), ci(c2) + 1)]


def extract(wb):
    d = wb[DASH]
    kpis = dict(
        vpl=round(num(d["J5"].value), 2),
        taxaVpl=round(num(d["K5"].value), 4),
        payback=int(num(d["L5"].value)),
        capitalGiro=round(num(d["O5"].value), 2),
        investimento=round(num(d["R5"].value), 2),
        necessidade=round(num(d["V5"].value), 2),
        tirAm=TIR_AM,
        tirAa=round(num(d["X5"].value), 10),
    )
    faturamento = [round(num(v), 2) for v in row(d, 65, "BA", "BE")]
    # Na planilha: BA76 = lucratividade (fração) | BA77 = lucro líquido (R$)
    lucratividade = [round(num(v) * 100, 2) for v in row(d, 76, "BA", "BE")]
    lucro = [round(num(v), 2) for v in row(d, 77, "BA", "BE")]
    rentabilidade = [round(num(v) * 100, 2) for v in row(d, 94, "BA", "BE")]
    fluxo = [round(num(v), 2) for v in row(d, 87, "BA", "DH")]

    # Investimento
    wi = wb["Investimento Inicial"]
    invest = []
    for r in range(4, 80):
        b = wi.cell(r, 2).value
        e = wi.cell(r, 5).value
        if b and isinstance(e, (int, float)) and e > 0:
            invest.append([esc(b), round(e, 2)])
    inv_total = round(sum(i[1] for i in invest), 2)
    invest = [[lab, v, round(v / inv_total * 100, 2)] for lab, v in invest]

    # Despesas Fixas
    wd = wb["Despesas Fixas"]
    despesas = []
    for r in range(7, 40):
        b = wd.cell(r, 2).value
        if not (b and str(b).strip()):
            continue
        anos = [round(num(wd.cell(r, c).value), 2) for c in range(4, 9)]
        if any(anos):
            despesas.append([esc(b), anos])
    desp_total = [round(num(wd.cell(40, c).value), 2) for c in range(4, 9)]

    # Recursos Humanos
    wr = wb["Recursos Humanos"]
    cargos = []
    for r in range(9, 16):
        b = wr.cell(r, 2).value
        if b and str(b).strip():
            cargos.append([esc(b), num(wr.cell(r, 3).value), round(num(wr.cell(r, 4).value), 2),
                           round(num(wr.cell(r, 11).value), 2)])
    rh_total, rh_qtd = [], []
    for r in range(1, 90):
        b = wr.cell(r, 2).value
        if b and "TOTAL MENSAL" in str(b):
            rh_total.append(round(num(wr.cell(r, 11).value), 2))
            rh_qtd.append(int(num(wr.cell(r, 3).value)))

    # DRE anual (soma 12 meses/ano; mês 1 = coluna H = idx 8)
    wdre = wb["DRE"]

    def annual(rr):
        return [round(sum(num(wdre.cell(rr, 8 + y * 12 + m).value) for m in range(12)), 0) for y in range(5)]

    dre_rows = [
        ("(+) Receita Bruta", 5), ("(-) Impostos", 30), ("(=) Receita Líquida", 33),
        ("(-) Folha de Pagamento", 49), ("(-) Despesas Fixas", 85),
        ("(-) Taxas do Sistema", 90), ("(-) Despesas Operacionais", 96),
        ("(=) Lucro Operacional", 100),
    ]
    dre_anual = [[lab, annual(rr)] for lab, rr in dre_rows]
    fat_bruto = annual(5)

    return dict(
        kpis=kpis, faturamento=faturamento, lucro=lucro, lucratividade=lucratividade,
        rentabilidade=rentabilidade, fluxo=fluxo, invest=invest, inv_total=inv_total,
        despesas=despesas, desp_total=desp_total, cargos=cargos, rh_total=rh_total,
        rh_qtd=rh_qtd, dre_anual=dre_anual, fat_bruto=fat_bruto,
    )


def emit_projeto(slug, nome, x):
    k = x["kpis"]
    L = []
    L.append("  {")
    L.append(f'    slug: "{slug}", nome: "{nome}",')
    L.append("    kpis: {")
    L.append(f"      vpl: {k['vpl']}, taxaVpl: {k['taxaVpl']}, payback: {k['payback']},")
    L.append(f"      capitalGiro: {k['capitalGiro']}, investimento: {k['investimento']},")
    L.append(f"      necessidade: {k['necessidade']}, tirAm: {k['tirAm']}, tirAa: {k['tirAa']},")
    L.append("    },")
    L.append(f"    faturamento: {arr(x['faturamento'])},")
    L.append(f"    lucro: {arr(x['lucro'])},")
    L.append(f"    lucratividade: {arr(x['lucratividade'])},")
    L.append(f"    rentabilidade: {arr(x['rentabilidade'])},")
    L.append(f"    fluxoCaixaAcumulado: {arr(x['fluxo'])},")
    L.append("    investimento: [")
    for lab, v, p in x["invest"]:
        L.append(f'      {{ label: "{lab}", valor: {v}, pct: {p} }},')
    L.append("    ],")
    L.append(f"    investimentoTotal: {x['inv_total']},")
    L.append("    despesasFixas: [")
    for lab, anos in x["despesas"]:
        L.append(f'      {{ label: "{lab}", anos: {arr(anos)} }},')
    L.append("    ],")
    L.append(f"    despesasFixasTotal: {arr(x['desp_total'])},")
    L.append("    rhCargos: [")
    for c, q, s, t in x["cargos"]:
        L.append(f'      {{ cargo: "{c}", qtd: {q}, salario: {s}, total: {t} }},')
    L.append("    ],")
    L.append(f"    rhTotalMensal: {arr(x['rh_total'])},")
    L.append(f"    rhQtd: {arr(x['rh_qtd'], 0)},")
    L.append("    dreAnual: [")
    for lab, anos in x["dre_anual"]:
        L.append(f'      {{ label: "{lab}", anos: {arr(anos, 0)} }},')
    L.append("    ],")
    L.append(f"    faturamentoBrutoAnual: {arr(x['fat_bruto'], 0)},")
    L.append("  },")
    return "\n".join(L)


def main():
    blocks = []
    for slug, nome, pid in PROJETOS:
        print(f"  baixando {nome} …")
        x = extract(fetch(pid))
        blocks.append(emit_projeto(slug, nome, x))
        print(f"    VPL={x['kpis']['vpl']:.0f}  Payback={x['kpis']['payback']}m  Invest={x['kpis']['investimento']:.0f}")

    head = '''// Dados de Viabilidade Financeira do Franqueado — TODOS os projetos.
// NÃO editar à mão — regenerar com `python scripts/gen_projetos.py`.

export interface Kpis {
  vpl: number; taxaVpl: number; payback: number; capitalGiro: number;
  investimento: number; necessidade: number; tirAm: number; tirAa: number;
}
export interface InvestItem { label: string; valor: number; pct: number; }
export interface SerieAno { label: string; anos: number[]; }
export interface CargoRH { cargo: string; qtd: number; salario: number; total: number; }
export interface Projeto {
  slug: string; nome: string; kpis: Kpis;
  faturamento: number[]; lucro: number[]; lucratividade: number[]; rentabilidade: number[];
  fluxoCaixaAcumulado: number[];
  investimento: InvestItem[]; investimentoTotal: number;
  despesasFixas: SerieAno[]; despesasFixasTotal: number[];
  rhCargos: CargoRH[]; rhTotalMensal: number[]; rhQtd: number[];
  dreAnual: SerieAno[]; faturamentoBrutoAnual: number[];
}

export const ANOS = ["Ano 1", "Ano 2", "Ano 3", "Ano 4", "Ano 5"] as const;

export const PROJETOS: Projeto[] = [
'''
    ts = head + "\n".join(blocks) + "\n];\n"
    out = os.path.abspath(OUT)
    with open(out, "w", encoding="utf-8", newline="\n") as fp:
        fp.write(ts)
    print(f"[OK] Gerado: {out}  ({len(PROJETOS)} projetos)")


if __name__ == "__main__":
    main()
