#!/usr/bin/env python3
"""
gen_detalhe.py — regenera web/lib/viabilidadeDetalhe.ts a partir da planilha publicada.

Consolida o detalhamento por aba do modelo de viabilidade (Projeto Compact 272K):
Investimento Inicial, Despesas Fixas, Recursos Humanos, DRE (anual) e Faturamento.
Para DRE soma os 12 meses de cada ano (colunas mensais a partir de H = mês 1).

Uso:
    pip install openpyxl
    python scripts/gen_detalhe.py
"""
import io
import os
import urllib.request

import openpyxl

PUB = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRxgJwgkxtlfhF5jXUgDw5go-APDt61uiKo3t-SeXDT2plBnfviczb-RpHkL2NQiA/pub?output=xlsx"
)
OUT = os.path.join(os.path.dirname(__file__), "..", "web", "lib", "viabilidadeDetalhe.ts")


def num(v):
    return v if isinstance(v, (int, float)) else 0.0


def load():
    with urllib.request.urlopen(PUB) as r:
        return openpyxl.load_workbook(io.BytesIO(r.read()), data_only=True)


def esc(s):
    return str(s).strip().replace("\\", "").replace('"', "'")


def arr(vals, nd=2):
    return "[" + ", ".join(str(round(v, nd)) for v in vals) + "]"


def main():
    wb = load()

    # ── Investimento Inicial ──────────────────────────────────────────────
    ws = wb["Investimento Inicial"]
    invest = []
    for r in range(4, 80):
        b = ws.cell(r, 2).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        if b and isinstance(e, (int, float)) and e > 0:
            invest.append((esc(b), round(e, 2), round((f or 0) * 100, 2)))
    invest_total = round(sum(i[1] for i in invest), 2)

    # ── Despesas Fixas (itens não-zero × 5 anos) + total ──────────────────
    df = wb["Despesas Fixas"]
    despesas = []
    for r in range(7, 40):
        b = df.cell(r, 2).value
        if not (b and str(b).strip()):
            continue
        anos = [round(num(df.cell(r, c).value), 2) for c in range(4, 9)]
        if any(anos):
            despesas.append((esc(b), anos))
    despesas_total = [round(num(df.cell(40, c).value), 2) for c in range(4, 9)]

    # ── Recursos Humanos ──────────────────────────────────────────────────
    rh = wb["Recursos Humanos"]
    cargos = []
    for r in range(9, 16):
        b = rh.cell(r, 2).value
        if b and str(b).strip():
            cargos.append((esc(b), num(rh.cell(r, 3).value), round(num(rh.cell(r, 4).value), 2),
                           round(num(rh.cell(r, 11).value), 2)))
    rh_total, rh_qtd = [], []
    for r in range(1, 90):
        b = rh.cell(r, 2).value
        if b and "TOTAL MENSAL" in str(b):
            rh_total.append(round(num(rh.cell(r, 11).value), 2))
            rh_qtd.append(num(rh.cell(r, 3).value))

    # ── DRE anual (soma 12 meses/ano, mês 1 = coluna H = idx 8) ───────────
    dre = wb["DRE"]

    def annual(row):
        return [round(sum(num(dre.cell(row, 8 + y * 12 + m).value) for m in range(12)), 2) for y in range(5)]

    dre_rows = [
        ("(+) Receita Bruta", 5), ("(-) Impostos", 30), ("(=) Receita Líquida", 33),
        ("(-) Folha de Pagamento", 49), ("(-) Despesas Fixas", 85),
        ("(-) Taxas do Sistema", 90), ("(-) Despesas Operacionais", 96),
        ("(=) Lucro Operacional", 100),
    ]
    dre_anual = [(lab, annual(r)) for lab, r in dre_rows]
    faturamento_bruto = annual(5)

    # ── Emite TS ──────────────────────────────────────────────────────────
    L = []
    L.append("// Detalhamento do modelo de Viabilidade — Projeto Compact 272K (abas da planilha).")
    L.append("// NÃO editar à mão — regenerar com `python scripts/gen_detalhe.py`.")
    L.append("")
    L.append("export interface InvestItem { label: string; valor: number; pct: number; }")
    L.append("export interface SerieAno { label: string; anos: number[]; }")
    L.append("export interface CargoRH { cargo: string; qtd: number; salario: number; total: number; }")
    L.append("")
    L.append("export const investimento: InvestItem[] = [")
    for lab, v, p in invest:
        L.append(f'  {{ label: "{lab}", valor: {v}, pct: {p} }},')
    L.append("];")
    L.append(f"export const investimentoTotal = {invest_total};")
    L.append("")
    L.append("export const despesasFixas: SerieAno[] = [")
    for lab, anos in despesas:
        L.append(f'  {{ label: "{lab}", anos: {arr(anos)} }},')
    L.append("];")
    L.append(f"export const despesasFixasTotal = {arr(despesas_total)};")
    L.append("")
    L.append("export const rhCargos: CargoRH[] = [")
    for c, q, s, t in cargos:
        L.append(f'  {{ cargo: "{c}", qtd: {q}, salario: {s}, total: {t} }},')
    L.append("];")
    L.append(f"export const rhTotalMensal = {arr(rh_total)};")
    L.append(f"export const rhQtd = {arr(rh_qtd, 0)};")
    L.append("")
    L.append("export const dreAnual: SerieAno[] = [")
    for lab, anos in dre_anual:
        L.append(f'  {{ label: "{lab}", anos: {arr(anos, 0)} }},')
    L.append("];")
    L.append(f"export const faturamentoBrutoAnual = {arr(faturamento_bruto, 0)};")
    L.append("")
    L.append("export const temFinanciamento = false;")
    L.append("")

    out = os.path.abspath(OUT)
    with open(out, "w", encoding="utf-8", newline="\n") as fp:
        fp.write("\n".join(L))
    print(f"[OK] Gerado: {out}")
    print(f"     invest={invest_total:.0f}  despesas/ano={despesas_total}  rh={rh_total}")


if __name__ == "__main__":
    main()
