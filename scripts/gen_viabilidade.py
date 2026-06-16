#!/usr/bin/env python3
"""
gen_viabilidade.py — regenera web/lib/viabilidade.ts a partir da planilha publicada.

A aba "Dashboard Franqueado (5anos)" do modelo de viabilidade (Projeto Compact 272K)
não é exposta pelo link CSV (que traz só 1 aba), mas o mesmo link de publicação,
com ?output=xlsx, devolve o workbook inteiro. Este script baixa o xlsx, lê os
valores calculados (data_only) e reescreve o snapshot TypeScript.

Uso:
    pip install openpyxl requests
    python scripts/gen_viabilidade.py

Ranges-fonte (aba "Dashboard Franqueado (5anos)"):
    KPIs        J5 (VPL), K5 (Taxa VPL), L5 (Payback), O5 (Capital de giro),
                R5 (Investimento), V5 (Necessidade), X5 (TIR a.a.)
    TIR a.m.    aba "Premissas" (3,44%) — fixo abaixo; ajuste se a planilha mudar.
    Séries (colunas ocultas, BA..):
                Faturamento    BA65:BE65
                Lucratividade% BA76:BE76   (fração → *100)
                Lucro          BA77:BE77
                Rentabilidade% BA94:BE94   (fração → *100)
                Fluxo Acum.    BA87:DH87   (60 colunas)
"""
import io
import os
import urllib.request

import openpyxl
from openpyxl.utils import column_index_from_string as ci

PUB_BASE = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRxgJwgkxtlfhF5jXUgDw5go-APDt61uiKo3t-SeXDT2plBnfviczb-RpHkL2NQiA/pub"
)
SHEET = "Dashboard Franqueado (5anos)"
OUT = os.path.join(os.path.dirname(__file__), "..", "web", "lib", "viabilidade.ts")
TIR_AM = 0.0344  # aba Premissas; ajuste se a planilha mudar


def fetch_ws():
    with urllib.request.urlopen(f"{PUB_BASE}?output=xlsx") as r:
        data = r.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return wb[SHEET]


def row(ws, r, c1, c2):
    return [ws.cell(r, c).value for c in range(ci(c1), ci(c2) + 1)]


def js_arr(vals, nd=2):
    return "[" + ", ".join(f"{round(v, nd)}" for v in vals) + "]"


def main():
    ws = fetch_ws()
    vpl = ws["J5"].value
    taxa_vpl = ws["K5"].value
    payback = ws["L5"].value
    capital = ws["O5"].value
    invest = ws["R5"].value
    necess = ws["V5"].value
    tir_aa = ws["X5"].value

    faturamento = row(ws, 65, "BA", "BE")
    lucrat = [v * 100 for v in row(ws, 76, "BA", "BE")]
    lucro = row(ws, 77, "BA", "BE")
    rent = [v * 100 for v in row(ws, 94, "BA", "BE")]
    fluxo = row(ws, 87, "BA", "DH")

    fluxo_lines = []
    for i in range(0, len(fluxo), 7):
        chunk = ", ".join(f"{round(v, 2)}" for v in fluxo[i : i + 7])
        fluxo_lines.append("  " + chunk + ",")
    fluxo_block = "\n".join(fluxo_lines)

    ts = f"""// Snapshot do modelo de Viabilidade Financeira do Franqueado — Projeto Compact 272K.
// Fonte: Google Sheets publicado, aba "Dashboard Franqueado (5anos)".
// NÃO editar à mão — regenerar com `python scripts/gen_viabilidade.py` quando a
// planilha mudar (ver scripts/README.md). Ranges-fonte documentados no script.

export const PROJETO = "Projeto Compact 272K";

export interface ViabKpis {{
  vpl: number;          // R$
  taxaVpl: number;      // fração (0.12 = 12%)
  payback: number;      // meses
  capitalGiro: number;  // R$ (negativo = saída)
  investimento: number; // R$ (negativo = saída)
  necessidade: number;  // R$ (negativo = saída)
  tirAm: number;        // fração ao mês
  tirAa: number;        // fração ao ano
}}

export const kpis: ViabKpis = {{
  vpl: {round(vpl, 2)},
  taxaVpl: {round(taxa_vpl, 4)},
  payback: {int(payback)},
  capitalGiro: {round(capital, 2)},
  investimento: {round(invest, 2)},
  necessidade: {round(necess, 2)},
  tirAm: {TIR_AM},
  tirAa: {round(tir_aa, 10)},
}};

export const anos = ["Ano 1", "Ano 2", "Ano 3", "Ano 4", "Ano 5"] as const;

// Faturamento bruto médio mensal por ano (R$)
export const faturamento = {js_arr(faturamento)};

// Lucro líquido médio mensal por ano (R$)
export const lucro = {js_arr(lucro)};

// Lucratividade por ano (% — lucro / faturamento)
export const lucratividade = {js_arr(lucrat)};

// Rentabilidade por ano (% — retorno sobre o capital investido)
export const rentabilidade = {js_arr(rent)};

// Fluxo de Caixa Acumulado — {len(fluxo)} pontos mensais (R$)
export const fluxoCaixaAcumulado = [
{fluxo_block}
];
"""
    out_path = os.path.abspath(OUT)
    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(ts)
    print(f"[OK] Gerado: {out_path}")
    print(f"     VPL={vpl:.0f}  TIR={tir_aa:.4f} a.a.  Payback={int(payback)}m  fluxo={len(fluxo)} pts")


if __name__ == "__main__":
    main()
